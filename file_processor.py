import os
import re
import json
import aiofiles
from typing import List, Dict, Optional
from pathlib import Path
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import openpyxl
from docx import Document
from PyPDF2 import PdfReader
from config import Config

# Инициализация EasyOCR (один раз при импорте) - опционально
reader = None
try:
    import easyocr
    reader = easyocr.Reader(['ru', 'en'], gpu=False)
except ImportError:
    print("EasyOCR не установлен, будет использован Tesseract")
    reader = None
except Exception as e:
    print(f"EasyOCR не доступен, будет использован Tesseract: {e}")
    reader = None

class FileProcessor:
    """Класс для обработки различных типов файлов и извлечения текста"""
    
    def __init__(self):
        self.upload_dir = Path(Config.UPLOAD_DIR)
        self.temp_dir = Path(Config.TEMP_DIR)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        self.temp_dir.mkdir(parents=True, exist_ok=True)
    
    async def save_file(self, file_data: bytes, filename: str) -> str:
        """Сохранение файла на диск"""
        file_path = self.upload_dir / filename
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(file_data)
        return str(file_path)
    
    async def extract_text_from_image(self, image_path: str) -> str:
        """Извлечение текста из изображения с помощью OCR"""
        try:
            # Используем EasyOCR если доступен, иначе Tesseract
            if reader:
                result = reader.readtext(image_path)
                text = "\n".join([item[1] for item in result])
            else:
                # Настраиваем путь к Tesseract
                if Config.TESSERACT_CMD:
                    pytesseract.pytesseract.tesseract_cmd = Config.TESSERACT_CMD
                image = Image.open(image_path)
                text = pytesseract.image_to_string(image, lang=Config.OCR_LANGUAGE)
            return text
        except Exception as e:
            print(f"Ошибка при извлечении текста из изображения: {e}")
            import traceback
            traceback.print_exc()
            return ""
    
    async def extract_text_from_pdf(self, pdf_path: str) -> str:
        """Извлечение текста из PDF"""
        text_parts = []
        
        try:
            # Метод 1: Прямое чтение текста из PDF
            reader_pdf = PdfReader(pdf_path)
            for page in reader_pdf.pages:
                text_parts.append(page.extract_text())
        except Exception as e:
            print(f"Ошибка при чтении PDF напрямую: {e}")
        
        # Метод 2: Конвертация в изображения и OCR (если прямой метод не сработал)
        if not any(text_parts):
            try:
                images = convert_from_path(pdf_path)
                for i, image in enumerate(images):
                    temp_image_path = self.temp_dir / f"pdf_page_{i}.png"
                    image.save(temp_image_path, "PNG")
                    ocr_text = await self.extract_text_from_image(str(temp_image_path))
                    text_parts.append(ocr_text)
                    temp_image_path.unlink()  # Удаляем временный файл
            except Exception as e:
                print(f"Ошибка при OCR PDF: {e}")
        
        return "\n".join(text_parts)
    
    async def extract_text_from_excel(self, excel_path: str) -> str:
        """Извлечение текста из Excel файла"""
        try:
            workbook = openpyxl.load_workbook(excel_path)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)
            
            return "\n".join(text_parts)
        except Exception as e:
            print(f"Ошибка при чтении Excel: {e}")
            return ""
    
    async def extract_text_from_word(self, word_path: str) -> str:
        """Извлечение текста из Word документа"""
        try:
            doc = Document(word_path)
            paragraphs = [para.text for para in doc.paragraphs]
            return "\n".join(paragraphs)
        except Exception as e:
            print(f"Ошибка при чтении Word: {e}")
            return ""
    
    async def process_file(self, file_path: str, file_type: str) -> str:
        """Обработка файла и извлечение текста в зависимости от типа"""
        file_path_obj = Path(file_path)
        
        if not file_path_obj.exists():
            raise FileNotFoundError(f"Файл не найден: {file_path}")
        
        if file_type.startswith("image/"):
            return await self.extract_text_from_image(file_path)
        elif file_type == "application/pdf":
            return await self.extract_text_from_pdf(file_path)
        elif file_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel"
        ]:
            return await self.extract_text_from_excel(file_path)
        elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return await self.extract_text_from_word(file_path)
        else:
            raise ValueError(f"Неподдерживаемый тип файла: {file_type}")
    
    def extract_article_numbers(self, text: str, articles: List[str]) -> List[Dict]:
        """Извлечение артикулов из текста и сопоставление с базой"""
        matches = []
        text_lower = text.lower()
        
        for article in articles:
            article_lower = article.lower().strip()
            if not article_lower:
                continue
            
            # Поиск точного совпадения
            if article_lower in text_lower:
                # Находим контекст вокруг артикула
                index = text_lower.find(article_lower)
                start = max(0, index - 50)
                end = min(len(text), index + len(article_lower) + 50)
                context = text[start:end]
                
                matches.append({
                    "article": article,
                    "found_text": context,
                    "confidence": 1.0,
                    "match_type": "exact"
                })
            
            # Поиск частичного совпадения (если артикул содержит цифры)
            elif any(char.isdigit() for char in article_lower):
                # Ищем паттерны, похожие на артикул
                pattern = re.escape(article_lower)
                pattern = pattern.replace("\\ ", "\\s*")
                regex = re.compile(pattern, re.IGNORECASE)
                
                for match in regex.finditer(text):
                    start = max(0, match.start() - 50)
                    end = min(len(text), match.end() + 50)
                    context = text[start:end]
                    
                    matches.append({
                        "article": article,
                        "found_text": context,
                        "confidence": 0.8,
                        "match_type": "partial"
                    })
        
        # Удаляем дубликаты
        seen = set()
        unique_matches = []
        for match in matches:
            key = (match["article"], match["found_text"][:100])
            if key not in seen:
                seen.add(key)
                unique_matches.append(match)
        
        return unique_matches

