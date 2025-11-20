from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional, Dict
from pydantic import BaseModel
from datetime import datetime
import json
import openpyxl
from io import BytesIO
import tempfile
import os
import uuid

from database import get_db, Article, ProcessedFile, MatchedArticle, ProductMapping, ConfirmedMapping, init_db
from file_processor import FileProcessor
from config import Config
from difflib import SequenceMatcher
import openai

app = FastAPI(title="Article Matcher API", version="1.0.0")

# CORS настройки
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # В продакшене указать конкретные домены
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

file_processor = FileProcessor()

# Pydantic модели
class ArticleCreate(BaseModel):
    article_number: str
    name: str
    description: Optional[str] = None
    price: Optional[float] = None
    category: Optional[str] = None

class ArticleResponse(BaseModel):
    id: int
    article_number: str
    name: str
    description: Optional[str]
    price: Optional[float]
    category: Optional[str]
    created_at: datetime
    
    class Config:
        from_attributes = True

class MatchedArticleResponse(BaseModel):
    id: int
    article_id: int
    article_number: str
    article_name: str
    found_text: str
    confidence: float
    created_at: datetime
    
    class Config:
        from_attributes = True

class ProcessedFileResponse(BaseModel):
    id: int
    user_id: int
    file_name: str
    file_type: str
    status: str
    matched_articles: List[MatchedArticleResponse]
    created_at: datetime
    
    class Config:
        from_attributes = True

# API Endpoints

@app.on_event("startup")
async def startup_event():
    """Инициализация при запуске"""
    await init_db()

@app.get("/")
async def root():
    """Корневой endpoint"""
    return {"message": "Article Matcher API", "version": "1.0.0"}

@app.get("/api/articles", response_model=List[ArticleResponse])
async def get_articles(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """Получение списка артикулов"""
    query = select(Article)
    
    if search:
        query = query.where(
            Article.article_number.contains(search) |
            Article.name.contains(search)
        )
    
    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    articles = result.scalars().all()
    return articles

@app.post("/api/articles", response_model=ArticleResponse)
async def create_article(
    article: ArticleCreate,
    db: AsyncSession = Depends(get_db)
):
    """Создание нового артикула"""
    # Проверка на дубликат
    existing = await db.execute(
        select(Article).where(Article.article_number == article.article_number)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Артикул уже существует")
    
    db_article = Article(**article.dict())
    db.add(db_article)
    await db.commit()
    await db.refresh(db_article)
    return db_article

@app.get("/api/articles/{article_id}", response_model=ArticleResponse)
async def get_article(article_id: int, db: AsyncSession = Depends(get_db)):
    """Получение артикула по ID"""
    result = await db.execute(select(Article).where(Article.id == article_id))
    article = result.scalar_one_or_none()
    if not article:
        raise HTTPException(status_code=404, detail="Артикул не найден")
    return article

@app.delete("/api/articles/{article_id}")
async def delete_article(article_id: int, db: AsyncSession = Depends(get_db)):
    """Удаление артикула"""
    result = await db.execute(select(Article).where(Article.id == article_id))
    article = result.scalar_one_or_none()
    if not article:
        raise HTTPException(status_code=404, detail="Артикул не найден")
    
    await db.delete(article)
    await db.commit()
    return {"message": "Артикул удален"}

@app.post("/api/files/upload")
async def upload_file(
    file: UploadFile = File(...),
    user_id: int = Query(...),
    db: AsyncSession = Depends(get_db)
):
    """Загрузка и обработка файла"""
    try:
        # Проверка типа файла
        if file.content_type not in Config.SUPPORTED_IMAGE_TYPES + Config.SUPPORTED_DOCUMENT_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"Неподдерживаемый тип файла: {file.content_type}"
            )
        
        # Чтение файла
        file_bytes = await file.read()
        
        if len(file_bytes) > Config.MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="Файл слишком большой")
        
        # Сохранение файла
        file_path = await file_processor.save_file(file_bytes, file.filename)
        
        # Создание записи в БД
        processed_file = ProcessedFile(
            user_id=user_id,
            file_name=file.filename,
            file_type=file.content_type,
            file_path=file_path,
            status="processing"
        )
        db.add(processed_file)
        await db.commit()
        await db.refresh(processed_file)
        
        # Извлечение текста
        extracted_text = await file_processor.process_file(file_path, file.content_type)
        
        # Получение артикулов из базы
        result = await db.execute(select(Article))
        articles = result.scalars().all()
        article_numbers = [article.article_number for article in articles]
        
        # Поиск совпадений
        matches = file_processor.extract_article_numbers(extracted_text, article_numbers)
        
        # Обновление записи
        processed_file.extracted_text = extracted_text[:10000]
        processed_file.matched_articles = json.dumps(matches, ensure_ascii=False)
        processed_file.status = "completed"
        
        # Сохранение деталей совпадений
        for match in matches:
            article_result = await db.execute(
                select(Article).where(Article.article_number == match["article"])
            )
            article = article_result.scalar_one_or_none()
            
            if article:
                matched_article = MatchedArticle(
                    processed_file_id=processed_file.id,
                    article_id=article.id,
                    found_text=match["found_text"],
                    confidence=match["confidence"]
                )
                db.add(matched_article)
        
        await db.commit()
        await db.refresh(processed_file)
        
        return {
            "id": processed_file.id,
            "status": "completed",
            "matches_count": len(matches),
            "matches": matches[:20]  # Первые 20 совпадений
        }
        
    except Exception as e:
        # Обновляем статус на ошибку
        if 'processed_file' in locals():
            processed_file.status = "error"
            await db.commit()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/files", response_model=List[ProcessedFileResponse])
async def get_files(
    user_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db)
):
    """Получение списка обработанных файлов"""
    query = select(ProcessedFile)
    
    if user_id:
        query = query.where(ProcessedFile.user_id == user_id)
    
    query = query.order_by(ProcessedFile.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    files = result.scalars().all()
    
    # Загружаем совпадения для каждого файла
    response = []
    for file in files:
        matches_result = await db.execute(
            select(MatchedArticle, Article)
            .join(Article, MatchedArticle.article_id == Article.id)
            .where(MatchedArticle.processed_file_id == file.id)
        )
        matches = matches_result.all()
        
        matched_articles = []
        for matched, article in matches:
            matched_articles.append(MatchedArticleResponse(
                id=matched.id,
                article_id=article.id,
                article_number=article.article_number,
                article_name=article.name,
                found_text=matched.found_text,
                confidence=matched.confidence,
                created_at=matched.created_at
            ))
        
        response.append(ProcessedFileResponse(
            id=file.id,
            user_id=file.user_id,
            file_name=file.file_name,
            file_type=file.file_type,
            status=file.status,
            matched_articles=matched_articles,
            created_at=file.created_at
        ))
    
    return response

@app.get("/api/files/{file_id}", response_model=ProcessedFileResponse)
async def get_file(file_id: int, db: AsyncSession = Depends(get_db)):
    """Получение деталей обработанного файла"""
    result = await db.execute(select(ProcessedFile).where(ProcessedFile.id == file_id))
    file = result.scalar_one_or_none()
    
    if not file:
        raise HTTPException(status_code=404, detail="Файл не найден")
    
    # Загружаем совпадения
    matches_result = await db.execute(
        select(MatchedArticle, Article)
        .join(Article, MatchedArticle.article_id == Article.id)
        .where(MatchedArticle.processed_file_id == file_id)
    )
    matches = matches_result.all()
    
    matched_articles = []
    for matched, article in matches:
        matched_articles.append(MatchedArticleResponse(
            id=matched.id,
            article_id=article.id,
            article_number=article.article_number,
            article_name=article.name,
            found_text=matched.found_text,
            confidence=matched.confidence,
            created_at=matched.created_at
        ))
    
    return ProcessedFileResponse(
        id=file.id,
        user_id=file.user_id,
        file_name=file.file_name,
        file_type=file.file_type,
        status=file.status,
        matched_articles=matched_articles,
        created_at=file.created_at
    )

@app.get("/api/stats")
async def get_stats(db: AsyncSession = Depends(get_db)):
    """Получение статистики"""
    articles_count = await db.scalar(select(func.count(Article.id)))
    files_count = await db.scalar(select(func.count(ProcessedFile.id)))
    matches_count = await db.scalar(select(func.count(MatchedArticle.id)))
    
    return {
        "articles_count": articles_count,
        "files_count": files_count,
        "matches_count": matches_count
    }

# ========== API для работы с таблицей сопоставления ==========

class ProductMappingCreate(BaseModel):
    # Старые поля (для совместимости)
    code_1c: Optional[str] = None
    bortlanger: Optional[str] = None
    epiroc: Optional[str] = None
    almazgeobur: Optional[str] = None
    competitors: Optional[dict] = None
    
    # Новые базовые поля
    article_bl: Optional[str] = None  # артикул bl
    article_agb: Optional[str] = None  # артикул агб
    variant_1: Optional[str] = None  # вариант подбора 1
    variant_2: Optional[str] = None  # вариант подбора 2
    variant_3: Optional[str] = None  # вариант подбора 3
    variant_4: Optional[str] = None  # вариант подбора 4
    variant_5: Optional[str] = None  # вариант подбора 5
    variant_6: Optional[str] = None  # вариант подбора 6
    variant_7: Optional[str] = None  # вариант подбора 7
    variant_8: Optional[str] = None  # вариант подбора 8
    unit: Optional[str] = None  # ед.изм.
    code: Optional[str] = None  # код
    nomenclature_agb: Optional[str] = None  # номенклатура агб
    packaging: Optional[str] = None  # фасовка для химии, кг.

class ProductMappingResponse(BaseModel):
    id: int
    # Старые поля (для совместимости)
    code_1c: Optional[str]
    bortlanger: Optional[str]
    epiroc: Optional[str]
    almazgeobur: Optional[str]
    competitors: Optional[dict]
    
    # Новые базовые поля
    article_bl: Optional[str]
    article_agb: Optional[str]
    variant_1: Optional[str]
    variant_2: Optional[str]
    variant_3: Optional[str]
    variant_4: Optional[str]
    variant_5: Optional[str]
    variant_6: Optional[str]
    variant_7: Optional[str]
    variant_8: Optional[str]
    unit: Optional[str]
    code: Optional[str]
    nomenclature_agb: Optional[str]
    packaging: Optional[str]
    
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class ProductMappingSearchResponse(BaseModel):
    mapping: ProductMappingResponse
    match_score: float
    matched_fields: List[str]

async def ai_analyze_excel_structure(file_path: str) -> Optional[Dict]:
    """Использует AI для анализа структуры Excel файла и определения столбцов"""
    if not Config.OPENAI_API_KEY:
        return None
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        if not workbook.sheetnames:
            return None
        
        # Берем первый лист
        sheet = workbook[workbook.sheetnames[0]]
        
        # Собираем первые 10 строк для анализа
        sample_data = []
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
            row_data = [str(cell) if cell is not None else '' for cell in row]
            sample_data.append({
                'row': row_idx,
                'values': row_data[:20]  # Первые 20 столбцов
            })
        
        # Формируем промпт для AI
        prompt = f"""Ты помощник для анализа структуры Excel таблицы.

Данные из файла (первые 10 строк):
{json.dumps(sample_data, ensure_ascii=False, indent=2)}

Задача: определи структуру таблицы и найди столбец, который содержит:
1. Артикулы или номера товаров (приоритет)
2. Если артикула нет - номенклатуру/название товара

ВАЖНО: 
- Столбцы нумеруются с 1 (первый столбец = 1)
- Если в первой строке есть заголовки - используй их для определения
- Верни номер столбца (начиная с 1), который содержит артикул/номер или номенклатуру

Ответь ТОЛЬКО в формате JSON:
{{
    "article_column": <номер столбца с артикулом/номером или null>,
    "nomenclature_column": <номер столбца с номенклатурой/названием или null>,
    "header_row": <номер строки с заголовками (обычно 1) или null>,
    "reasoning": "<краткое объяснение>"
}}"""

        # Вызываем OpenAI API
        client = openai.OpenAI(api_key=Config.OPENAI_API_KEY)
        response = client.chat.completions.create(
            model=Config.OPENAI_MODEL,
            messages=[
                {"role": "system", "content": "Ты помощник для анализа структуры таблиц. Отвечай только в формате JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=300
        )
        
        ai_response = response.choices[0].message.content.strip()
        
        # Парсим ответ AI
        try:
            if ai_response.startswith("```"):
                ai_response = ai_response.split("```")[1]
                if ai_response.startswith("json"):
                    ai_response = ai_response[4:]
            ai_response = ai_response.strip()
            
            ai_result = json.loads(ai_response)
            return ai_result
        except json.JSONDecodeError:
            return None
    except Exception as e:
        print(f"Ошибка AI-анализа структуры: {e}")
        return None

async def ai_interpret_text(recognized_text: str, available_mappings: List[ProductMapping], db: AsyncSession) -> Optional[Dict]:
    """Использует AI для интерпретации распознанного текста и поиска в БД"""
    if not Config.OPENAI_API_KEY:
        return None  # Если нет API ключа, возвращаем None
    
    try:
        # Сначала проверяем подтвержденные сопоставления
        confirmed_result = await db.execute(
            select(ConfirmedMapping).where(
                ConfirmedMapping.recognized_text.ilike(f"%{recognized_text}%")
            )
        )
        confirmed = confirmed_result.scalar_one_or_none()
        if confirmed:
            # Находим mapping по ID
            mapping_result = await db.execute(
                select(ProductMapping).where(ProductMapping.id == confirmed.mapping_id)
            )
            mapping = mapping_result.scalar_one_or_none()
            if mapping:
                return {
                    'mapping_id': mapping.id,
                    'match_score': 100.0,  # Подтвержденные сопоставления имеют 100%
                    'is_confirmed': True,
                    'mapping': mapping
                }
        
        # Подготавливаем данные для AI (первые 30 записей для экономии токенов)
        # Приоритет записям с артикулом АГБ
        mappings_with_agb = [m for m in available_mappings if m.article_agb][:30]
        if not mappings_with_agb:
            mappings_with_agb = available_mappings[:30]
        
        sample_data = []
        for m in mappings_with_agb:
            sample_data.append({
                'id': m.id,
                'article_agb': m.article_agb or '',
                'nomenclature_agb': m.nomenclature_agb or '',
                'code': m.code or '',
                'variant_1': m.variant_1 or '',
                'variant_2': m.variant_2 or '',
            })
        
        # Формируем промпт для AI
        prompt = f"""Ты помощник для поиска соответствий в базе данных товаров.

Распознанный текст из файла: "{recognized_text}"

Доступные записи в базе данных (примеры, всего записей в БД: {len(available_mappings)}):
{json.dumps(sample_data, ensure_ascii=False, indent=2)}

Задача: найди наиболее подходящую запись из базы данных для распознанного текста.
Учитывай возможные опечатки, сокращения и вариации написания.
Ищи совпадения по артикулам, кодам, номенклатуре и вариантам подбора.

ВАЖНО: Если в базе данных есть запись с точно таким же текстом или очень похожим (более 80% совпадения), верни её ID.
Если не уверен (менее 50% совпадения), верни mapping_id: null.

Ответь ТОЛЬКО в формате JSON:
{{
    "mapping_id": <id записи из базы или null>,
    "confidence": <0-100, процент уверенности>,
    "reasoning": "<краткое объяснение почему выбрана эта запись>"
}}

Если не нашел подходящей записи, верни mapping_id: null."""

        # Вызываем OpenAI API
        client = openai.OpenAI(api_key=Config.OPENAI_API_KEY)
        response = client.chat.completions.create(
            model=Config.OPENAI_MODEL,
            messages=[
                {"role": "system", "content": "Ты помощник для поиска соответствий в базе данных. Отвечай только в формате JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,  # Низкая температура для более точных результатов
            max_tokens=200
        )
        
        ai_response = response.choices[0].message.content.strip()
        
        # Парсим ответ AI
        try:
            # Убираем markdown код блоки если есть
            if ai_response.startswith("```"):
                ai_response = ai_response.split("```")[1]
                if ai_response.startswith("json"):
                    ai_response = ai_response[4:]
            ai_response = ai_response.strip()
            
            ai_result = json.loads(ai_response)
            
            if ai_result.get('mapping_id') and ai_result.get('confidence', 0) > 50:
                # Находим mapping по ID
                mapping_result = await db.execute(
                    select(ProductMapping).where(ProductMapping.id == ai_result['mapping_id'])
                )
                mapping = mapping_result.scalar_one_or_none()
                if mapping:
                    return {
                        'mapping_id': mapping.id,
                        'match_score': float(ai_result.get('confidence', 0)),
                        'is_ai_match': True,
                        'reasoning': ai_result.get('reasoning', ''),
                        'mapping': mapping
                    }
        except json.JSONDecodeError:
            pass  # Если не удалось распарсить, возвращаем None
        
        return None
    except Exception as e:
        print(f"Ошибка AI-поиска: {e}")
        return None  # В случае ошибки возвращаем None, будет использован обычный поиск

def calculate_similarity(text1: str, text2: str) -> float:
    """Вычисляет процент совпадения между двумя строками на основе совпадения слов"""
    if not text1 or not text2:
        return 0.0
    
    # Нормализуем тексты: приводим к нижнему регистру и разбиваем на слова
    import re
    words1 = set(re.findall(r'\w+', text1.lower()))
    words2 = set(re.findall(r'\w+', text2.lower()))
    
    if not words1 or not words2:
        return 0.0
    
    # Находим точные совпадения слов
    exact_matches = words1.intersection(words2)
    
    # Если есть точные совпадения, считаем процент на их основе
    if exact_matches:
        # Процент = (количество совпавших слов / количество слов в запросе) * 100
        # Но также учитываем, сколько слов из текста совпало
        match_ratio = len(exact_matches) / len(words1)
        # Дополнительный бонус, если все слова запроса найдены
        if len(exact_matches) == len(words1):
            match_ratio = 1.0
        return match_ratio * 100
    
    # Если нет точных совпадений, используем частичное совпадение
    # Проверяем, содержит ли текст2 слова из text1 (частичное совпадение)
    partial_score = 0.0
    for word1 in words1:
        for word2 in words2:
            if word1 in word2 or word2 in word1:
                partial_score += 0.5  # Частичное совпадение дает 50% от полного
                break
    
    if partial_score > 0:
        return (partial_score / len(words1)) * 100
    
    # Если нет даже частичных совпадений, используем SequenceMatcher как fallback
    return SequenceMatcher(None, text1.lower(), text2.lower()).ratio() * 100

def normalize_field(value: Optional[str]) -> Optional[str]:
    """Нормализует поле: пустые значения и "-" становятся None"""
    if not value or value.strip() == '' or value.strip() == '-':
        return None
    return value.strip()

@app.post("/api/mappings", response_model=ProductMappingResponse)
async def create_mapping(mapping: ProductMappingCreate, db: AsyncSession = Depends(get_db)):
    """Создание новой строки в таблице сопоставления"""
    db_mapping = ProductMapping(
        # Старые поля
        code_1c=normalize_field(mapping.code_1c),
        bortlanger=normalize_field(mapping.bortlanger),
        epiroc=normalize_field(mapping.epiroc),
        almazgeobur=normalize_field(mapping.almazgeobur),
        competitors=mapping.competitors or {},
        # Новые базовые поля
        article_bl=normalize_field(mapping.article_bl),
        article_agb=normalize_field(mapping.article_agb),
        variant_1=normalize_field(mapping.variant_1),
        variant_2=normalize_field(mapping.variant_2),
        variant_3=normalize_field(mapping.variant_3),
        variant_4=normalize_field(mapping.variant_4),
        variant_5=normalize_field(mapping.variant_5),
        variant_6=normalize_field(mapping.variant_6),
        variant_7=normalize_field(mapping.variant_7),
        variant_8=normalize_field(mapping.variant_8),
        unit=normalize_field(mapping.unit),
        code=normalize_field(mapping.code),
        nomenclature_agb=normalize_field(mapping.nomenclature_agb),
        packaging=normalize_field(mapping.packaging)
    )
    db.add(db_mapping)
    await db.commit()
    await db.refresh(db_mapping)
    return db_mapping

@app.get("/api/mappings")
async def get_mappings(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db)
):
    """Получение всех строк таблицы с пагинацией"""
    try:
        # Получаем общее количество записей
        count_result = await db.execute(select(func.count(ProductMapping.id)))
        total = count_result.scalar() or 0
        
        # Получаем данные с пагинацией
        result = await db.execute(
            select(ProductMapping)
            .order_by(ProductMapping.id)
            .offset(skip)
            .limit(limit)
        )
        mappings = result.scalars().all()
        
        return {
            "items": [ProductMappingResponse.model_validate(m) for m in mappings],
            "total": total,
            "skip": skip,
            "limit": limit
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при загрузке таблицы: {str(e)}")

@app.get("/api/mappings/search", response_model=List[ProductMappingSearchResponse])
async def search_mappings(
    query: str = Query(..., description="Поисковый запрос"),
    min_score: float = Query(50.0, description="Минимальный процент совпадения"),
    limit: int = Query(20, description="Максимальное количество результатов"),
    db: AsyncSession = Depends(get_db)
):
    """Поиск строк с процентом совпадения на основе совпадения слов"""
    try:
        if not query or not query.strip():
            return []
        
        result = await db.execute(select(ProductMapping))
        all_mappings = result.scalars().all()
        
        search_results = []
        
        for mapping in all_mappings:
            scores = []
            matched_fields = []
            
            # Проверяем все поля (старые и новые)
            fields_to_check = {
                'code_1c': mapping.code_1c,
                'bortlanger': mapping.bortlanger,
                'epiroc': mapping.epiroc,
                'almazgeobur': mapping.almazgeobur,
                # Новые базовые поля
                'article_bl': mapping.article_bl,
                'article_agb': mapping.article_agb,
                'variant_1': mapping.variant_1,
                'variant_2': mapping.variant_2,
                'variant_3': mapping.variant_3,
                'variant_4': mapping.variant_4,
                'variant_5': mapping.variant_5,
                'variant_6': mapping.variant_6,
                'variant_7': mapping.variant_7,
                'variant_8': mapping.variant_8,
                'unit': mapping.unit,
                'code': mapping.code,
                'nomenclature_agb': mapping.nomenclature_agb,
                'packaging': mapping.packaging,
            }
            
            # Проверяем конкурентов
            if mapping.competitors:
                for comp_name, comp_value in mapping.competitors.items():
                    if comp_value:
                        fields_to_check[comp_name] = comp_value
            
            for field_name, field_value in fields_to_check.items():
                if field_value:
                    score = calculate_similarity(query, str(field_value))
                    if score > 0:
                        scores.append(score)
                        if score >= min_score:
                            matched_fields.append(field_name)
            
            if scores:
                max_score = max(scores)
                if max_score >= min_score:
                    search_results.append({
                        'mapping': mapping,
                        'match_score': round(max_score, 2),
                        'matched_fields': matched_fields
                    })
    
        # Сортируем по проценту совпадения (по убыванию)
        search_results.sort(key=lambda x: x['match_score'], reverse=True)
        
        # Ограничиваем количество результатов
        search_results = search_results[:limit]
        
        return [
            ProductMappingSearchResponse(
                mapping=ProductMappingResponse.model_validate(item['mapping']),
                match_score=item['match_score'],
                matched_fields=item['matched_fields']
            )
            for item in search_results
        ]
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при поиске: {str(e)}")

@app.get("/api/mappings/{mapping_id}", response_model=ProductMappingResponse)
async def get_mapping(mapping_id: int, db: AsyncSession = Depends(get_db)):
    """Получение конкретной строки"""
    result = await db.execute(select(ProductMapping).where(ProductMapping.id == mapping_id))
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    return mapping

@app.put("/api/mappings/{mapping_id}", response_model=ProductMappingResponse)
async def update_mapping(
    mapping_id: int,
    mapping: ProductMappingCreate,
    db: AsyncSession = Depends(get_db)
):
    """Обновление строки таблицы"""
    result = await db.execute(select(ProductMapping).where(ProductMapping.id == mapping_id))
    db_mapping = result.scalar_one_or_none()
    if not db_mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    
    # Обновляем старые поля
    db_mapping.code_1c = normalize_field(mapping.code_1c)
    db_mapping.bortlanger = normalize_field(mapping.bortlanger)
    db_mapping.epiroc = normalize_field(mapping.epiroc)
    db_mapping.almazgeobur = normalize_field(mapping.almazgeobur)
    if mapping.competitors is not None:
        db_mapping.competitors = mapping.competitors
    
    # Обновляем новые базовые поля
    db_mapping.article_bl = normalize_field(mapping.article_bl)
    db_mapping.article_agb = normalize_field(mapping.article_agb)
    db_mapping.variant_1 = normalize_field(mapping.variant_1)
    db_mapping.variant_2 = normalize_field(mapping.variant_2)
    db_mapping.variant_3 = normalize_field(mapping.variant_3)
    db_mapping.variant_4 = normalize_field(mapping.variant_4)
    db_mapping.variant_5 = normalize_field(mapping.variant_5)
    db_mapping.variant_6 = normalize_field(mapping.variant_6)
    db_mapping.variant_7 = normalize_field(mapping.variant_7)
    db_mapping.variant_8 = normalize_field(mapping.variant_8)
    db_mapping.unit = normalize_field(mapping.unit)
    db_mapping.code = normalize_field(mapping.code)
    db_mapping.nomenclature_agb = normalize_field(mapping.nomenclature_agb)
    db_mapping.packaging = normalize_field(mapping.packaging)
    
    await db.commit()
    await db.refresh(db_mapping)
    return db_mapping

@app.delete("/api/mappings/{mapping_id}")
async def delete_mapping(mapping_id: int, db: AsyncSession = Depends(get_db)):
    """Удаление строки таблицы"""
    result = await db.execute(select(ProductMapping).where(ProductMapping.id == mapping_id))
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    await db.delete(mapping)
    await db.commit()
    return {"message": "Mapping deleted"}

@app.post("/api/mappings/upload")
async def upload_mapping_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """Загрузка файла с интеллектуальным распознаванием и сопоставлением с таблицей соответствий"""
    try:
        # Сохраняем файл
        file_bytes = await file.read()
        file_path = await file_processor.save_file(file_bytes, file.filename)
        
        # Получаем все записи из таблицы соответствий
        result = await db.execute(select(ProductMapping))
        all_mappings = result.scalars().all()
        
        # Результаты распознавания и сопоставления
        # Включаем все обработанные строки, даже если ничего не найдено
        recognition_results = []
        all_processed_items = []  # Все обработанные элементы для отчета
        lines = None  # Инициализируем переменную для не-Excel файлов
        
        # Если это Excel файл - используем интеллектуальный анализ структуры
        if file.content_type in [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
            'application/octet-stream'
        ] or (file.filename and file.filename.lower().endswith(('.xlsx', '.xls', '.csv'))):
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                
                # Анализируем структуру файла с помощью AI
                structure = await ai_analyze_excel_structure(file_path)
                
                # Обрабатываем каждый лист
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Определяем столбцы для поиска
                    article_col = None
                    nomenclature_col = None
                    header_row = 1
                    
                    if structure:
                        article_col = structure.get('article_column')
                        nomenclature_col = structure.get('nomenclature_column')
                        header_row = structure.get('header_row', 1)
                    
                    # Если AI не определил, пробуем найти по заголовкам
                    if not article_col and not nomenclature_col:
                        # Ищем заголовки в первых строках
                        for row_idx in range(1, min(5, sheet.max_row + 1)):
                            row = sheet[row_idx]
                            for col_idx, cell in enumerate(row, start=1):
                                cell_value = str(cell.value).lower() if cell.value else ''
                                if any(keyword in cell_value for keyword in ['артикул', 'номер', 'код', 'article', 'number']):
                                    article_col = col_idx
                                elif any(keyword in cell_value for keyword in ['номенклатура', 'название', 'наименование', 'name', 'nomenclature']):
                                    nomenclature_col = col_idx
                            if article_col or nomenclature_col:
                                header_row = row_idx
                                break
                    
                    # Обрабатываем строки данных
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                        # Извлекаем значение из нужного столбца
                        search_value = None
                        
                        if article_col:
                            col_idx = article_col - 1  # openpyxl использует 0-based индексы для values_only
                            if col_idx < len(row) and row[col_idx]:
                                search_value = str(row[col_idx]).strip()
                        
                        if not search_value and nomenclature_col:
                            col_idx = nomenclature_col - 1
                            if col_idx < len(row) and row[col_idx]:
                                search_value = str(row[col_idx]).strip()
                        
                        if not search_value or len(search_value) < 2:
                            continue
                        
                        # Сохраняем что искали
                        processed_item = {
                            'recognized_text': search_value,
                            'mapping_id': None,
                            'match_score': None,
                            'matched_field': None,
                            'matched_value': None,
                            'mapping': None,
                            'is_ai_match': False,
                            'is_confirmed': False
                        }
                        
                        # Ищем совпадения
                        best_match = None
                        best_score = 0.0
                        
                        # Сначала пробуем AI-поиск
                        ai_match = await ai_interpret_text(search_value, all_mappings, db)
                        if ai_match and ai_match.get('match_score', 0) > best_score:
                            best_score = ai_match['match_score']
                            mapping = ai_match['mapping']
                            best_match = {
                                'recognized_text': search_value,
                                'mapping_id': mapping.id,
                                'match_score': round(best_score, 2),
                                'matched_field': 'ai_match',
                                'matched_value': search_value,
                                'is_ai_match': True,
                                'is_confirmed': ai_match.get('is_confirmed', False),
                                'mapping': {
                                    'id': mapping.id,
                                    'article_bl': mapping.article_bl,
                                    'article_agb': mapping.article_agb,
                                    'variant_1': mapping.variant_1,
                                    'variant_2': mapping.variant_2,
                                    'variant_3': mapping.variant_3,
                                    'variant_4': mapping.variant_4,
                                    'variant_5': mapping.variant_5,
                                    'variant_6': mapping.variant_6,
                                    'variant_7': mapping.variant_7,
                                    'variant_8': mapping.variant_8,
                                    'unit': mapping.unit,
                                    'code': mapping.code,
                                    'nomenclature_agb': mapping.nomenclature_agb,
                                    'packaging': mapping.packaging,
                                }
                            }
                        
                        # Если AI не нашел или результат слабый, используем обычный поиск
                        if not best_match or best_score < 80:
                            for mapping in all_mappings:
                                fields_to_check = [
                                    ('article_bl', mapping.article_bl),
                                    ('article_agb', mapping.article_agb),
                                    ('variant_1', mapping.variant_1),
                                    ('variant_2', mapping.variant_2),
                                    ('variant_3', mapping.variant_3),
                                    ('variant_4', mapping.variant_4),
                                    ('variant_5', mapping.variant_5),
                                    ('variant_6', mapping.variant_6),
                                    ('variant_7', mapping.variant_7),
                                    ('variant_8', mapping.variant_8),
                                    ('code', mapping.code),
                                    ('nomenclature_agb', mapping.nomenclature_agb),
                                ]
                                
                                for field_name, field_value in fields_to_check:
                                    if field_value:
                                        score = calculate_similarity(search_value, str(field_value))
                                        if score > best_score:
                                            best_score = score
                                            best_match = {
                                                'recognized_text': search_value,
                                                'mapping_id': mapping.id,
                                                'match_score': round(score, 2),
                                                'matched_field': field_name,
                                                'matched_value': field_value,
                                                'mapping': {
                                                    'id': mapping.id,
                                                    'article_bl': mapping.article_bl,
                                                    'article_agb': mapping.article_agb,
                                                    'variant_1': mapping.variant_1,
                                                    'variant_2': mapping.variant_2,
                                                    'variant_3': mapping.variant_3,
                                                    'variant_4': mapping.variant_4,
                                                    'variant_5': mapping.variant_5,
                                                    'variant_6': mapping.variant_6,
                                                    'variant_7': mapping.variant_7,
                                                    'variant_8': mapping.variant_8,
                                                    'unit': mapping.unit,
                                                    'code': mapping.code,
                                                    'nomenclature_agb': mapping.nomenclature_agb,
                                                    'packaging': mapping.packaging,
                                                }
                                            }
                        
                        # Обновляем processed_item с результатом поиска
                        if best_match:
                            processed_item.update(best_match)
                            recognition_results.append(best_match)
                        else:
                            # Если ничего не найдено, все равно добавляем в список
                            processed_item['matched_value'] = 'Не найдено'
                        
                        # Добавляем в общий список всех обработанных элементов
                        all_processed_items.append(processed_item)
                
            except Exception as e:
                print(f"Ошибка при обработке Excel: {e}")
                import traceback
                traceback.print_exc()
                # Fallback на обычную обработку
                extracted_text = await file_processor.process_file(file_path, file.content_type)
                if not extracted_text or not extracted_text.strip():
                    raise HTTPException(status_code=400, detail="Не удалось извлечь текст из файла")
                
                lines = [line.strip() for line in extracted_text.split('\n') if line.strip()]
        else:
            # Для других типов файлов используем обычную обработку
            extracted_text = await file_processor.process_file(file_path, file.content_type)
            
            if not extracted_text or not extracted_text.strip():
                raise HTTPException(status_code=400, detail="Не удалось извлечь текст из файла")
            
            # Разбиваем текст на строки (артикулы/названия)
            lines = [line.strip() for line in extracted_text.split('\n') if line.strip()]
        
        # Для не-Excel файлов обрабатываем построчно
        # Проверяем, были ли обработаны строки из Excel
        excel_processed = len(all_processed_items) > 0
        
        if not excel_processed and lines:
            for line in lines:
                if not line or len(line) < 2:
                    continue
                
                # Сохраняем что искали
                processed_item = {
                    'recognized_text': line,
                    'mapping_id': None,
                    'match_score': None,
                    'matched_field': None,
                    'matched_value': None,
                    'mapping': None,
                    'is_ai_match': False,
                    'is_confirmed': False
                }
                
                best_match = None
                best_score = 0.0
                
                # Сначала пробуем AI-поиск
                ai_match = await ai_interpret_text(line, all_mappings, db)
                if ai_match and ai_match.get('match_score', 0) > best_score:
                    best_score = ai_match['match_score']
                    mapping = ai_match['mapping']
                    best_match = {
                        'recognized_text': line,
                        'mapping_id': mapping.id,
                        'match_score': round(best_score, 2),
                        'matched_field': 'ai_match',
                        'matched_value': line,
                        'is_ai_match': True,
                        'is_confirmed': ai_match.get('is_confirmed', False),
                        'mapping': {
                            'id': mapping.id,
                            'article_bl': mapping.article_bl,
                            'article_agb': mapping.article_agb,
                            'variant_1': mapping.variant_1,
                            'variant_2': mapping.variant_2,
                            'variant_3': mapping.variant_3,
                            'variant_4': mapping.variant_4,
                            'variant_5': mapping.variant_5,
                            'variant_6': mapping.variant_6,
                            'variant_7': mapping.variant_7,
                            'variant_8': mapping.variant_8,
                            'unit': mapping.unit,
                            'code': mapping.code,
                            'nomenclature_agb': mapping.nomenclature_agb,
                            'packaging': mapping.packaging,
                        }
                    }
                
                # Если AI не нашел или результат слабый, используем обычный поиск
                if not best_match or best_score < 80:
                    # Ищем совпадения во всех полях таблицы
                    for mapping in all_mappings:
                        # Проверяем все поля
                        fields_to_check = [
                            ('article_bl', mapping.article_bl),
                            ('article_agb', mapping.article_agb),
                            ('variant_1', mapping.variant_1),
                            ('variant_2', mapping.variant_2),
                            ('variant_3', mapping.variant_3),
                            ('variant_4', mapping.variant_4),
                            ('variant_5', mapping.variant_5),
                            ('variant_6', mapping.variant_6),
                            ('variant_7', mapping.variant_7),
                            ('variant_8', mapping.variant_8),
                            ('code', mapping.code),
                            ('nomenclature_agb', mapping.nomenclature_agb),
                        ]
                        
                        for field_name, field_value in fields_to_check:
                            if field_value:
                                score = calculate_similarity(line, str(field_value))
                                if score > best_score:
                                    best_score = score
                                    best_match = {
                                        'recognized_text': line,
                                        'mapping_id': mapping.id,
                                        'match_score': round(score, 2),
                                        'matched_field': field_name,
                                        'matched_value': field_value,
                                        'mapping': {
                                            'id': mapping.id,
                                            'article_bl': mapping.article_bl,
                                            'article_agb': mapping.article_agb,
                                            'variant_1': mapping.variant_1,
                                            'variant_2': mapping.variant_2,
                                            'variant_3': mapping.variant_3,
                                            'variant_4': mapping.variant_4,
                                            'variant_5': mapping.variant_5,
                                            'variant_6': mapping.variant_6,
                                            'variant_7': mapping.variant_7,
                                            'variant_8': mapping.variant_8,
                                            'unit': mapping.unit,
                                            'code': mapping.code,
                                            'nomenclature_agb': mapping.nomenclature_agb,
                                            'packaging': mapping.packaging,
                                        }
                                    }
                
                # Обновляем processed_item с результатом поиска
                if best_match:
                    processed_item.update(best_match)
                    recognition_results.append(best_match)
                else:
                    # Если ничего не найдено, все равно добавляем в список
                    processed_item['matched_value'] = 'Не найдено'
                
                # Добавляем в общий список всех обработанных элементов
                all_processed_items.append(processed_item)
        
        # Сохраняем результаты в сессию (можно использовать Redis или БД)
        # Пока сохраняем в файл временно
        recognized_count = len(all_processed_items)
        
        if recognized_count == 0:
            raise HTTPException(status_code=400, detail="Не удалось обработать файл. Убедитесь, что файл содержит данные.")
        
        session_id = str(uuid.uuid4())
        results_file = os.path.join(Config.TEMP_DIR, f"results_{session_id}.json")
        os.makedirs(Config.TEMP_DIR, exist_ok=True)
        try:
            with open(results_file, 'w', encoding='utf-8') as f:
                json.dump(all_processed_items, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка при сохранении результатов: {e}")
            # Продолжаем работу даже если не удалось сохранить в файл
        
        return {
            "message": f"Обработано {recognized_count} строк, найдено {len(recognition_results)} совпадений",
            "recognized_count": recognized_count,
            "matches_count": len(recognition_results),
            "results": all_processed_items,  # Возвращаем все результаты, включая "не найдено"
            "session_id": session_id
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при обработке файла: {str(e)}")

@app.post("/api/mappings/confirm")
async def confirm_mapping(
    recognized_text: str = Query(..., description="Распознанный текст"),
    mapping_id: int = Query(..., description="ID сопоставления"),
    match_score: float = Query(None, description="Процент совпадения"),
    db: AsyncSession = Depends(get_db)
):
    """Подтверждение сопоставления для использования в будущем"""
    try:
        # Проверяем, существует ли уже такое подтверждение
        existing = await db.execute(
            select(ConfirmedMapping).where(
                ConfirmedMapping.recognized_text == recognized_text,
                ConfirmedMapping.mapping_id == mapping_id
            )
        )
        confirmed = existing.scalar_one_or_none()
        
        if confirmed:
            # Увеличиваем счетчик подтверждений
            confirmed.user_confirmed += 1
            confirmed.match_score = match_score or confirmed.match_score
            confirmed.updated_at = datetime.utcnow()
        else:
            # Создаем новое подтверждение
            confirmed = ConfirmedMapping(
                recognized_text=recognized_text,
                mapping_id=mapping_id,
                match_score=match_score or 100.0,
                user_confirmed=1
            )
            db.add(confirmed)
        
        await db.commit()
        await db.refresh(confirmed)
        
        return {
            "message": "Сопоставление подтверждено",
            "confirmed_id": confirmed.id,
            "user_confirmed": confirmed.user_confirmed
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при подтверждении: {str(e)}")

@app.get("/api/mappings/upload/export/{session_id}")
async def export_recognition_results(session_id: str, background_tasks: BackgroundTasks):
    """Выгрузка результатов распознавания в Excel"""
    try:
        results_file = os.path.join(Config.TEMP_DIR, f"results_{session_id}.json")
        
        if not os.path.exists(results_file):
            raise HTTPException(status_code=404, detail="Результаты не найдены")
        
        # Загружаем результаты
        with open(results_file, 'r', encoding='utf-8') as f:
            results = json.load(f)
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты распознавания"
        
        # Заголовки
        headers = [
            'Распознанный текст',
            'ID соответствия',
            'Процент совпадения',
            'Поле совпадения',
            'Значение совпадения',
            'Артикул BL',
            'Артикул АГБ',
            'Вариант подбора 1',
            'Вариант подбора 2',
            'Вариант подбора 3',
            'Вариант подбора 4',
            'Вариант подбора 5',
            'Вариант подбора 6',
            'Вариант подбора 7',
            'Вариант подбора 8',
            'Ед.изм.',
            'Код',
            'Номенклатура АГБ',
            'Фасовка для химии, кг.'
        ]
        
        ws.append(headers)
        
        # Данные
        for result in results:
            mapping = result.get('mapping') or {}
            # Обрабатываем случаи, когда mapping может быть None или пустым
            if not mapping:
                mapping = {}
            
            # Обрабатываем match_score - может быть None
            match_score = result.get('match_score')
            if match_score is None:
                match_score = 0
            
            row = [
                result.get('recognized_text', '') or '',
                result.get('mapping_id', '') or '',
                match_score,
                result.get('matched_field', '') or '',
                result.get('matched_value', '') or '',
                mapping.get('article_bl', '') or '',
                mapping.get('article_agb', '') or '',
                mapping.get('variant_1', '') or '',
                mapping.get('variant_2', '') or '',
                mapping.get('variant_3', '') or '',
                mapping.get('variant_4', '') or '',
                mapping.get('variant_5', '') or '',
                mapping.get('variant_6', '') or '',
                mapping.get('variant_7', '') or '',
                mapping.get('variant_8', '') or '',
                mapping.get('unit', '') or '',
                mapping.get('code', '') or '',
                mapping.get('nomenclature_agb', '') or '',
                mapping.get('packaging', '') or '',
            ]
            ws.append(row)
        
        # Сохраняем во временный файл
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=Config.TEMP_DIR)
        temp_file_path = temp_file.name
        temp_file.close()
        
        # Убеждаемся, что директория существует
        os.makedirs(Config.TEMP_DIR, exist_ok=True)
        
        wb.save(temp_file_path)
        
        # Используем FileResponse для отправки файла (проще и надежнее)
        # Добавляем задачу на удаление файла после отправки
        def cleanup_file():
            try:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
            except Exception as e:
                print(f"Ошибка при удалении временного файла: {e}")
        
        background_tasks.add_task(cleanup_file)
        
        # Создаем FileResponse с правильными заголовками для скачивания
        response = FileResponse(
            temp_file_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"results_{session_id}.xlsx",
            headers={
                'Content-Disposition': f'attachment; filename="results_{session_id}.xlsx"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при выгрузке: {str(e)}")

@app.post("/api/mappings/upload-confirmations")
async def upload_confirmations_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """Загрузка Excel файла с подтвержденными сопоставлениями для массового обновления"""
    try:
        # Проверяем, что это Excel файл
        if not (file.filename and file.filename.lower().endswith(('.xlsx', '.xls'))):
            raise HTTPException(status_code=400, detail="Поддерживаются только Excel файлы (.xlsx, .xls)")
        
        # Сохраняем файл
        file_bytes = await file.read()
        file_path = await file_processor.save_file(file_bytes, file.filename)
        
        # Загружаем Excel файл
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Ищем заголовки
        header_row = 1
        headers = {}
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            cell_value = str(cell.value).lower() if cell.value else ''
            if 'распознанный' in cell_value or 'текст' in cell_value or 'что искалось' in cell_value:
                headers['recognized_text'] = col_idx
            elif 'id' in cell_value and 'соответствия' in cell_value or 'mapping_id' in cell_value:
                headers['mapping_id'] = col_idx
            elif 'совпадение' in cell_value or 'процент' in cell_value or 'match_score' in cell_value:
                headers['match_score'] = col_idx
        
        if 'recognized_text' not in headers or 'mapping_id' not in headers:
            raise HTTPException(
                status_code=400, 
                detail="Файл должен содержать колонки: 'Распознанный текст' (или 'Что искалось') и 'ID соответствия'"
            )
        
        # Обрабатываем строки
        confirmed_count = 0
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            try:
                # Получаем значения из нужных колонок
                recognized_text_col = headers['recognized_text'] - 1
                mapping_id_col = headers['mapping_id'] - 1
                match_score_col = headers.get('match_score', 0) - 1 if 'match_score' in headers else None
                
                if recognized_text_col >= len(row) or mapping_id_col >= len(row):
                    continue
                
                recognized_text = str(row[recognized_text_col]).strip() if row[recognized_text_col] else None
                mapping_id_str = str(row[mapping_id_col]).strip() if row[mapping_id_col] else None
                match_score = None
                
                if match_score_col is not None and match_score_col >= 0 and match_score_col < len(row):
                    try:
                        match_score = float(row[match_score_col]) if row[match_score_col] else None
                    except (ValueError, TypeError):
                        match_score = None
                
                if not recognized_text or not mapping_id_str:
                    continue
                
                # Парсим mapping_id
                try:
                    mapping_id = int(float(mapping_id_str))
                except (ValueError, TypeError):
                    errors.append(f"Строка {row_idx}: неверный ID соответствия '{mapping_id_str}'")
                    continue
                
                # Проверяем, существует ли mapping
                mapping_result = await db.execute(
                    select(ProductMapping).where(ProductMapping.id == mapping_id)
                )
                mapping = mapping_result.scalar_one_or_none()
                
                if not mapping:
                    errors.append(f"Строка {row_idx}: сопоставление с ID {mapping_id} не найдено")
                    continue
                
                # Создаем или обновляем подтверждение
                existing = await db.execute(
                    select(ConfirmedMapping).where(
                        ConfirmedMapping.recognized_text == recognized_text,
                        ConfirmedMapping.mapping_id == mapping_id
                    )
                )
                confirmed = existing.scalar_one_or_none()
                
                if confirmed:
                    confirmed.user_confirmed += 1
                    if match_score is not None:
                        confirmed.match_score = match_score
                    confirmed.updated_at = datetime.utcnow()
                else:
                    confirmed = ConfirmedMapping(
                        recognized_text=recognized_text,
                        mapping_id=mapping_id,
                        match_score=match_score or 100.0,
                        user_confirmed=1
                    )
                    db.add(confirmed)
                
                confirmed_count += 1
                
            except Exception as e:
                errors.append(f"Строка {row_idx}: ошибка обработки - {str(e)}")
                continue
        
        await db.commit()
        
        return {
            "message": f"Обработано подтверждений: {confirmed_count}",
            "confirmed_count": confirmed_count,
            "errors": errors[:10] if errors else [],  # Первые 10 ошибок
            "errors_count": len(errors)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при обработке файла: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

