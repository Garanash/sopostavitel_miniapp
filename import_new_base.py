#!/usr/bin/env python3
"""
Скрипт для импорта данных из файла 'База данных соответствий.xlsx'
в таблицу product_mappings с новой структурой полей
"""

import asyncio
import sys
from pathlib import Path
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import delete
from database import ProductMapping, Base
from config import Config
import openpyxl

async def clear_mappings(session: AsyncSession):
    """Удаление всех записей из таблицы product_mappings"""
    print("Очистка таблицы product_mappings...")
    await session.execute(delete(ProductMapping))
    await session.commit()
    print("✅ Таблица очищена")

async def import_excel_data(session: AsyncSession, file_path: str):
    """Импорт данных из Excel файла"""
    print(f"Чтение файла: {file_path}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"Размер таблицы: {ws.max_row} строк, {ws.max_column} колонок")
    
    # Находим заголовки
    headers = []
    header_row = None
    
    for row_idx, row in enumerate(ws.iter_rows(max_row=20), start=1):
        row_values = [cell.value for cell in row if cell.value]
        if len(row_values) >= 2:
            # Проверяем, есть ли известные заголовки
            row_lower = [str(v).lower() if v else '' for v in row_values]
            row_text = ' '.join(row_lower)
            # Ищем новые базовые поля
            if any(keyword in row_text for keyword in ['артикул bl', 'артикул агб', 'вариант подбора', 'ед.изм', 'номенклатура', 'фасовка', 'код']):
                headers = row_values
                header_row = row_idx
                print(f"Найдены заголовки в строке {row_idx}: {headers[:15]}")
                break
    
    if not headers:
        # Если заголовки не найдены, используем первую строку
        headers = [cell.value for cell in ws[1] if cell.value]
        header_row = 1
        print(f"Используем первую строку как заголовки: {headers[:15]}")
    
    # Определяем индексы колонок для новых полей
    header_lower = [str(h).lower() if h else '' for h in headers]
    
    # Новые базовые поля
    article_bl_idx = None
    article_agb_idx = None
    variant_indices = {}  # {1: idx, 2: idx, ...}
    unit_idx = None
    code_idx = None
    nomenclature_agb_idx = None
    packaging_idx = None
    
    # Старые поля (для совместимости)
    code_1c_idx = None
    bortlanger_idx = None
    epiroc_idx = None
    almazgeobur_idx = None
    competitor_indices = {}
    
    for idx, header in enumerate(header_lower):
        header_str = str(header).strip()
        
        # Новые базовые поля
        if 'артикул bl' in header_str or 'артикул bl' in header_str:
            article_bl_idx = idx
        elif 'артикул агб' in header_str or 'артикул агб' in header_str:
            article_agb_idx = idx
        elif 'вариант подбора' in header_str:
            # Извлекаем номер варианта
            try:
                # Ищем число в строке
                import re
                numbers = re.findall(r'\d+', header_str)
                if numbers:
                    variant_num = int(numbers[0])
                    if 1 <= variant_num <= 8:
                        variant_indices[variant_num] = idx
            except:
                pass
        elif 'ед.изм' in header_str or 'единица' in header_str:
            unit_idx = idx
        elif ('код' in header_str and '1с' not in header_str and 'номенклатура' not in header_str) or header_str == 'код':
            code_idx = idx
        elif 'номенклатура агб' in header_str or 'номенклатура' in header_str:
            nomenclature_agb_idx = idx
        elif 'фасовка' in header_str or 'кг' in header_str:
            packaging_idx = idx
        # Старые поля
        elif 'код' in header_str and '1с' in header_str:
            code_1c_idx = idx
        elif 'bortlanger' in header_str or 'бортлангер' in header_str:
            bortlanger_idx = idx
        elif 'epiroc' in header_str or 'эпирок' in header_str:
            epiroc_idx = idx
        elif 'almazgeobur' in header_str or 'алмазгеобур' in header_str:
            almazgeobur_idx = idx
        else:
            # Остальные колонки - конкуренты (если не системные)
            if header_str and header_str not in ['id', 'действия', 'action']:
                competitor_indices[idx] = header_str
    
    print(f"\nИндексы колонок:")
    print(f"  article_bl: {article_bl_idx}")
    print(f"  article_agb: {article_agb_idx}")
    print(f"  variant_1-8: {variant_indices}")
    print(f"  unit: {unit_idx}")
    print(f"  code: {code_idx}")
    print(f"  nomenclature_agb: {nomenclature_agb_idx}")
    print(f"  packaging: {packaging_idx}")
    print(f"  конкурентов: {len(competitor_indices)}")
    
    # Импортируем данные
    imported = 0
    skipped = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        row_values = [cell.value for cell in row]
        
        # Пропускаем пустые строки
        if not any(row_values):
            continue
        
        # Извлекаем значения для новых полей
        def get_value(idx):
            if idx is not None and idx < len(row_values) and row_values[idx]:
                val = str(row_values[idx]).strip()
                if val and val.lower() not in ['none', '-', '', 'none']:
                    return val
            return None
        
        article_bl = get_value(article_bl_idx)
        article_agb = get_value(article_agb_idx)
        variant_1 = get_value(variant_indices.get(1))
        variant_2 = get_value(variant_indices.get(2))
        variant_3 = get_value(variant_indices.get(3))
        variant_4 = get_value(variant_indices.get(4))
        variant_5 = get_value(variant_indices.get(5))
        variant_6 = get_value(variant_indices.get(6))
        variant_7 = get_value(variant_indices.get(7))
        variant_8 = get_value(variant_indices.get(8))
        unit = get_value(unit_idx)
        code = get_value(code_idx)
        nomenclature_agb = get_value(nomenclature_agb_idx)
        packaging = get_value(packaging_idx)
        
        # Старые поля (для совместимости)
        code_1c = get_value(code_1c_idx)
        bortlanger = get_value(bortlanger_idx)
        epiroc = get_value(epiroc_idx)
        almazgeobur = get_value(almazgeobur_idx)
        
        # Собираем конкурентов
        competitors = {}
        for idx, name in competitor_indices.items():
            val = get_value(idx)
            if val:
                competitors[name] = val
        
        # Пропускаем полностью пустые строки
        if not any([article_bl, article_agb, variant_1, variant_2, variant_3, variant_4, 
                   variant_5, variant_6, variant_7, variant_8, unit, code, 
                   nomenclature_agb, packaging, code_1c, bortlanger, epiroc, almazgeobur, competitors]):
            skipped += 1
            continue
        
        # Создаем запись
        mapping = ProductMapping(
            # Новые базовые поля
            article_bl=article_bl,
            article_agb=article_agb,
            variant_1=variant_1,
            variant_2=variant_2,
            variant_3=variant_3,
            variant_4=variant_4,
            variant_5=variant_5,
            variant_6=variant_6,
            variant_7=variant_7,
            variant_8=variant_8,
            unit=unit,
            code=code,
            nomenclature_agb=nomenclature_agb,
            packaging=packaging,
            # Старые поля (для совместимости)
            code_1c=code_1c,
            bortlanger=bortlanger,
            epiroc=epiroc,
            almazgeobur=almazgeobur,
            competitors=competitors if competitors else None
        )
        
        session.add(mapping)
        imported += 1
        
        # Коммитим батчами
        if imported % 1000 == 0:
            await session.commit()
            print(f"Импортировано: {imported} записей...")
    
    await session.commit()
    print(f"\n✅ Импорт завершен!")
    print(f"   Импортировано: {imported} записей")
    print(f"   Пропущено: {skipped} пустых строк")

async def main():
    """Основная функция"""
    file_path = Path("База данных соответствий.xlsx")
    
    if not file_path.exists():
        print(f"❌ Файл не найден: {file_path}")
        sys.exit(1)
    
    # Подключаемся к БД
    engine = create_async_engine(Config.DATABASE_URL, echo=False)
    async_session_maker = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    
    async with async_session_maker() as session:
        # Очищаем таблицу
        await clear_mappings(session)
        
        # Импортируем данные
        await import_excel_data(session, str(file_path))
    
    await engine.dispose()
    print("\n✅ Готово!")

if __name__ == "__main__":
    asyncio.run(main())
