"""
Скрипт для импорта данных из файла "КП БИ+Химия.xlsx"
Обрабатывает 2 листа, добавляет варианты подбора без дублирования
"""
import asyncio
import openpyxl
from database import init_db, async_session_maker, ProductMapping
from config import Config

def normalize_value(value):
    """Нормализует значение: None, пустые строки и '-' становятся None"""
    if value is None:
        return None
    val = str(value).strip()
    if val == '' or val == '-' or val.lower() == 'none':
        return None
    return val

async def find_or_create_mapping(session, article_agb, nomenclature_agb, code, unit, packaging):
    """Находит существующую запись или создает новую по артикулу АГБ"""
    if not article_agb:
        return None
    
    # Ищем существующую запись по артикулу АГБ
    from sqlalchemy import select
    result = await session.execute(
        select(ProductMapping).where(ProductMapping.article_agb == article_agb)
    )
    existing = result.scalar_one_or_none()
    
    if existing:
        # Обновляем поля, если они пустые
        if not existing.nomenclature_agb and nomenclature_agb:
            existing.nomenclature_agb = nomenclature_agb
        if not existing.code and code:
            existing.code = code
        if not existing.unit and unit:
            existing.unit = unit
        if not existing.packaging and packaging:
            existing.packaging = packaging
        return existing
    
    # Создаем новую запись
    new_mapping = ProductMapping(
        article_bl=None,
        article_agb=article_agb,
        variant_1=None,
        variant_2=None,
        variant_3=None,
        variant_4=None,
        variant_5=None,
        variant_6=None,
        variant_7=None,
        variant_8=None,
        unit=unit,
        code=code,
        nomenclature_agb=nomenclature_agb,
        packaging=packaging
    )
    session.add(new_mapping)
    await session.flush()
    return new_mapping

def add_variant_to_mapping(mapping, variant_value):
    """Добавляет вариант подбора в первую свободную позицию (variant_1 - variant_8)"""
    if not variant_value:
        return
    
    variants = [
        mapping.variant_1,
        mapping.variant_2,
        mapping.variant_3,
        mapping.variant_4,
        mapping.variant_5,
        mapping.variant_6,
        mapping.variant_7,
        mapping.variant_8,
    ]
    
    # Проверяем, нет ли уже такого варианта
    if variant_value in variants:
        return
    
    # Находим первую свободную позицию
    if not mapping.variant_1:
        mapping.variant_1 = variant_value
    elif not mapping.variant_2:
        mapping.variant_2 = variant_value
    elif not mapping.variant_3:
        mapping.variant_3 = variant_value
    elif not mapping.variant_4:
        mapping.variant_4 = variant_value
    elif not mapping.variant_5:
        mapping.variant_5 = variant_value
    elif not mapping.variant_6:
        mapping.variant_6 = variant_value
    elif not mapping.variant_7:
        mapping.variant_7 = variant_value
    elif not mapping.variant_8:
        mapping.variant_8 = variant_value

async def import_sheet(session, sheet, sheet_name):
    """Импортирует данные из одного листа"""
    print(f"\n📋 Обработка листа: {sheet_name}")
    print(f"Размер: {sheet.max_row} строк, {sheet.max_column} колонок")
    
    # Ищем заголовки
    headers = []
    header_row = None
    
    for row_idx in range(1, min(20, sheet.max_row + 1)):
        row = sheet[row_idx]
        row_values = [cell.value for cell in row if cell.value]
        if len(row_values) >= 2:
            row_lower = [str(v).lower() if v else '' for v in row_values]
            row_text = ' '.join(row_lower)
            
            # Ищем ключевые слова
            if any(keyword in row_text for keyword in [
                'артикул агб', 'артикул', 'номенклатура', 'код', 
                'вариант', 'подбор', 'ед.изм', 'фасовка'
            ]):
                headers = [cell.value for cell in row]
                header_row = row_idx
                print(f"✅ Найдены заголовки в строке {row_idx}")
                break
    
    if not headers:
        headers = [cell.value for cell in sheet[1]]
        header_row = 1
        print(f"⚠️ Используем первую строку как заголовки")
    
    # Определяем индексы колонок
    header_lower = [str(h).lower() if h else '' for h in headers]
    
    article_agb_idx = None
    nomenclature_agb_idx = None
    code_idx = None
    unit_idx = None
    packaging_idx = None
    variant_indices = {}  # Словарь для вариантов подбора
    
    for idx, header in enumerate(header_lower):
        if 'артикул агб' in header or (header == 'артикул' and article_agb_idx is None):
            article_agb_idx = idx
        elif 'номенклатура' in header or 'наименование' in header:
            nomenclature_agb_idx = idx
        elif header == 'код':
            code_idx = idx
        elif 'ед.изм' in header or 'единица' in header:
            unit_idx = idx
        elif 'фасовка' in header or 'упаковка' in header:
            packaging_idx = idx
        elif 'вариант' in header or 'подбор' in header:
            # Извлекаем номер варианта из заголовка
            import re
            match = re.search(r'(\d+)', header)
            if match:
                variant_num = int(match.group(1))
                if 1 <= variant_num <= 8:
                    variant_indices[variant_num] = idx
            else:
                # Если номер не указан, добавляем в первую свободную позицию
                for v in range(1, 9):
                    if v not in variant_indices.values():
                        variant_indices[v] = idx
                        break
    
    print(f"📊 Найденные колонки:")
    print(f"  - Артикул АГБ: {article_agb_idx + 1 if article_agb_idx is not None else 'не найден'}")
    print(f"  - Номенклатура АГБ: {nomenclature_agb_idx + 1 if nomenclature_agb_idx is not None else 'не найден'}")
    print(f"  - Код: {code_idx + 1 if code_idx is not None else 'не найден'}")
    print(f"  - Ед.изм: {unit_idx + 1 if unit_idx is not None else 'не найден'}")
    print(f"  - Фасовка: {packaging_idx + 1 if packaging_idx is not None else 'не найден'}")
    print(f"  - Варианты подбора: {len(variant_indices)}")
    
    # Импортируем данные
    imported = 0
    updated = 0
    skipped = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # Извлекаем значения
        def get_value(idx):
            if idx is not None and idx < len(row) and row[idx]:
                return normalize_value(row[idx])
            return None
        
        article_agb = get_value(article_agb_idx)
        nomenclature_agb = get_value(nomenclature_agb_idx)
        code = get_value(code_idx)
        unit = get_value(unit_idx)
        packaging = get_value(packaging_idx)
        
        # Пропускаем строки без артикула АГБ
        if not article_agb:
            skipped += 1
            continue
        
        # Находим или создаем запись
        mapping = await find_or_create_mapping(
            session, article_agb, nomenclature_agb, code, unit, packaging
        )
        
        if not mapping:
            skipped += 1
            continue
        
        is_new = mapping.id is None or not hasattr(mapping, '_sa_instance_state') or mapping._sa_instance_state.pending
        
        # Добавляем варианты подбора
        for variant_num, col_idx in variant_indices.items():
            variant_value = get_value(col_idx)
            if variant_value:
                add_variant_to_mapping(mapping, variant_value)
        
        if is_new:
            imported += 1
        else:
            updated += 1
    
    await session.commit()
    print(f"✅ Лист '{sheet_name}': импортировано {imported}, обновлено {updated}, пропущено {skipped}")
    return imported, updated, skipped

async def import_kp_bi_himiya(file_path: str):
    """Импорт данных из файла КП БИ+Химия.xlsx"""
    await init_db()
    
    print(f"📂 Открываю файл: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print(f"📋 Найдено листов: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    total_imported = 0
    total_updated = 0
    total_skipped = 0
    
    async with async_session_maker() as session:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            imported, updated, skipped = await import_sheet(session, sheet, sheet_name)
            total_imported += imported
            total_updated += updated
            total_skipped += skipped
    
    print(f"\n✅ Импорт завершен!")
    print(f"📊 Итого: импортировано {total_imported}, обновлено {total_updated}, пропущено {total_skipped}")

if __name__ == "__main__":
    file_path = "КП БИ+Химия.xlsx"
    asyncio.run(import_kp_bi_himiya(file_path))

