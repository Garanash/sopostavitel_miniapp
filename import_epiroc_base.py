"""
Скрипт для импорта базы данных Epiroc из Excel файла
"""
import asyncio
import openpyxl
from database import init_db, async_session_maker, ProductMapping
from config import Config

async def import_epiroc_base(file_path: str):
    """Импорт базы данных Epiroc из Excel файла"""
    await init_db()
    
    wb = openpyxl.load_workbook(file_path)
    # Используем первый лист
    ws = wb[wb.sheetnames[0]]
    
    # Ищем строку с заголовками (может быть не первая)
    header_row = 1
    headers = []
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_values = [cell.value for cell in ws[row_idx]]
        # Проверяем, есть ли в строке ключевые слова
        row_text = ' '.join([str(v) for v in row_values if v]).lower()
        if any(keyword in row_text for keyword in ['код', '1с', 'epiroc', 'bortlanger', 'алмазгеобур']):
            header_row = row_idx
            headers = row_values
            break
    
    if not headers:
        # Используем первую строку
        headers = [cell.value if cell.value else f"col_{cell.column_letter}" for cell in ws[1]]
    
    print(f"Строка заголовков: {header_row}")
    print(f"Найдено колонок: {len(headers)}")
    print(f"Заголовки: {headers}")
    
    # Если файл имеет простую структуру (2 колонки), обрабатываем как список артикулов
    # Первая колонка - название/артикул, вторая - значение/количество
    if len(headers) == 2:
        print("\n⚠️ Обнаружена простая структура (2 колонки)")
        print("Обрабатываем как список артикулов Epiroc")
        
        # Определяем, какая колонка содержит артикулы
        # Обычно первая колонка - это артикул/название
        epiroc_col = 0  # Первая колонка - артикулы Epiroc
        value_col = 1 if len(headers) > 1 else None  # Вторая колонка - значение
        
        print(f"Колонка Epiroc (артикулы): {epiroc_col}")
        print(f"Колонка значений: {value_col}")
    
    # Импортируем данные (начинаем со строки после заголовков)
    async with async_session_maker() as session:
        imported_count = 0
        skipped_count = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            # Пропускаем пустые строки
            if not any(row):
                continue
            
            # Если простая структура (2 колонки)
            if len(headers) == 2:
                epiroc_value = str(row[0]).strip() if row[0] else None
                other_value = str(row[1]).strip() if len(row) > 1 and row[1] else None
                
                if not epiroc_value or epiroc_value == 'None':
                    skipped_count += 1
                    continue
                
                # Создаем запись с артикулом Epiroc
                # Второе значение добавляем как конкурента
                competitors = {}
                if other_value and other_value != 'None':
                    # Используем название первой колонки как имя конкурента
                    competitor_name = str(headers[1]) if headers[1] else "Значение"
                    competitors[competitor_name] = other_value
                
                mapping = ProductMapping(
                    code_1c=None,
                    bortlanger=None,
                    epiroc=epiroc_value,
                    almazgeobur=None,
                    competitors=competitors if competitors else None
                )
            else:
                # Стандартная обработка с несколькими колонками
                code_1c = str(row[code_1c_idx]).strip() if code_1c_idx is not None and code_1c_idx < len(row) and row[code_1c_idx] else None
                bortlanger = str(row[bortlanger_idx]).strip() if bortlanger_idx is not None and bortlanger_idx < len(row) and row[bortlanger_idx] else None
                epiroc = str(row[epiroc_idx]).strip() if epiroc_idx is not None and epiroc_idx < len(row) and row[epiroc_idx] else None
                almazgeobur = str(row[almazgeobur_idx]).strip() if almazgeobur_idx is not None and almazgeobur_idx < len(row) and row[almazgeobur_idx] else None
                
                # Проверяем, что есть хотя бы одно значение
                if not any([code_1c, bortlanger, epiroc, almazgeobur]):
                    skipped_count += 1
                    continue
                
                # Собираем конкурентов
                competitors = {}
                for idx, name in competitor_indices.items():
                    if idx < len(row) and row[idx]:
                        value = str(row[idx]).strip()
                        if value and value != 'None':
                            competitors[name] = value
                
                mapping = ProductMapping(
                    code_1c=code_1c if code_1c and code_1c != 'None' else None,
                    bortlanger=bortlanger if bortlanger and bortlanger != 'None' else None,
                    epiroc=epiroc if epiroc and epiroc != 'None' else None,
                    almazgeobur=almazgeobur if almazgeobur and almazgeobur != 'None' else None,
                    competitors=competitors if competitors else None
                )
            
            session.add(mapping)
            imported_count += 1
            
            # Коммитим каждые 100 записей
            if imported_count % 100 == 0:
                await session.commit()
                print(f"Импортировано {imported_count} записей...")
        
        # Финальный коммит
        await session.commit()
        
        print(f"\n✅ Импорт завершен!")
        print(f"   Импортировано: {imported_count} записей")
        print(f"   Пропущено: {skipped_count} пустых строк")

if __name__ == "__main__":
    file_path = "База по заказчика Epiroc.xlsx"
    asyncio.run(import_epiroc_base(file_path))

